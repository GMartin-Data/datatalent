"""Parsing du XLSX BMO: extraction, filtrage IT, normalisation."""

from pathlib import Path

from openpyxl import load_workbook

from ingestion.bmo.config import (
    BMO_SHEET_NAME,
    CODES_METIER_IT,
    COLUMN_MAPPING,
    NUMERIC_COLUMNS,
)


def _build_header_index(header_row: tuple) -> dict[str, int]:
    """Associe chaque nom de colonne source à son index positionnel.

    Seules les colonnes présentes dans COLUMN_MAPPING sont retenues.
    Lève ValueError si une colonne attendue est absente du XLSX -
    signal immédiat d'un changement de schéma chez France Travail.
    """
    # {valeur_cellule: position} pour toute la ligne d'en-tête
    available = {cell.value: idx for idx, cell in enumerate(header_row)}

    header_index: dict[str, int] = {}
    missing = []

    for source_col in COLUMN_MAPPING:
        if source_col in available:
            header_index[source_col] = available[source_col]
        else:
            missing.append(source_col)

    if missing:
        raise ValueError(
            f"Colonnes attendues absentes du XLSX: {missing}. "
            f"Colonnes disponibles: {sorted(available.keys())}"
        )

    return header_index


def _normalize_value(source_col: str, raw_value) -> int | str | None:
    """Normalise une valeur brute selon le type de colonne.

    - année: cast int (toujours présent, jamais "*")
    - met/xmet/smet: "*" → None (secret statistique), sinon cast int
    - tout le reste: str tel quel (typage fin en dbt staging)
    """
    if source_col == "annee":
        return int(raw_value)

    if source_col in NUMERIC_COLUMNS:
        if raw_value == "*":
            return None
        else:
            return int(raw_value)

    # Colonnes texte - str() défensif au cas où openpyxl
    # retournerait un type inattendu (int pour REG, etc.)
    return str(raw_value)


def parse_bmo_xlsx(xlsx_path: str | Path) -> list[dict]:
    """Parse le XLSX BMO et retourne les lignes IT normalisées.

    Args:
        xlsx_path: chemin du fichier XLSX sur disque

    Returns:
        Liste de dicts, un par ligne retenue. Clés = noms JSONL
        (ex: "code_metier_bmo", "projets_recrutement").
        Prêt pour json.dumps() ligne par ligne.
    """
    # read_only=True: openpyxl ne charge pas tout en mémoire
    # il streame les lignes. Important sur un fichier de 50k lignes
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[BMO_SHEET_NAME]

    rows = ws.iter_rows()

    # Première ligne = en-têtes
    header_row = next(rows)
    header_index = _build_header_index(header_row)

    # Index de la colonne de filtrage - résolu une seule fois
    code_metier_idx = header_index["Code métier BMO"]

    records: list[dict] = []

    for row in rows:
        # Filtrage précoce: on teste le code métier avant de lire
        # les autres colonnes. Évite du travail inutile sur les ~49k lignes non IT.
        code_metier_value = row[code_metier_idx].value
        if code_metier_value not in CODES_METIER_IT:
            continue

        record = {}
        for source_col, jsonl_col in COLUMN_MAPPING.items():
            raw_value = row[header_index[source_col]].value
            record[jsonl_col] = _normalize_value(source_col, raw_value)

        records.append(record)

    wb.close()

    return records
